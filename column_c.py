from flask import Flask, jsonify, request
import os
import openpyxl

app = Flask(__name__)

def read_column_c():
    # Get the current directory
    current_dir = os.path.dirname(os.path.abspath(__file__))

    # Construct the file path
    file_path = os.path.join(current_dir, 'Hapag Loyd - Sample Data.xlsx')

    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active
    column_c = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        value_c = row[2]
        if value_c is not None:
            column_c.append(value_c)
    return column_c

def read_column_b():
    # Get the current directory
    current_dir = os.path.dirname(os.path.abspath(__file__))

    # Construct the file path
    file_path = os.path.join(current_dir, 'Hapag Loyd - Sample Data.xlsx')

    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active
    column_b = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        value_b = row[1]
        if value_b is not None:
            column_b.append(value_b)
    return column_b

def generate_series_numbers(start, end):
    return [f"{num:08d}" for num in range(start, end + 1)]

@app.route('/read-column-c', methods=['GET'])
def api_read_column_c():
    column_c_data = read_column_c()
    return jsonify(column_c_data)

@app.route('/read-column-b', methods=['GET'])
def api_read_column_b():
    column_b_data = read_column_b()
    return jsonify(column_b_data)

@app.route('/generate-series-numbers', methods=['POST'])
def api_generate_series_numbers():
    data = request.get_json()
    count = data.get('count')
    
    if not count or not isinstance(count, int):
        return jsonify({"error": "Please provide a valid 'count' in the request body"}), 400
    
    start_number = 10000000
    end_number = start_number + count - 1
    series_numbers = generate_series_numbers(start_number, end_number)
    return jsonify(series_numbers)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)